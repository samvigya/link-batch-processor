import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import io
import zipfile
from datetime import datetime
import os

# Page configuration
st.set_page_config(
    page_title="Link Batch Processor",
    page_icon="🔗",
    layout="wide"
)

# Function to find template files
def find_template_file(platform):
    """Find template file in multiple possible locations"""
    if platform == "Instagram":
        possible_names = [
            "IG_Influencers_100.xlsx",
            "/mnt/user-data/uploads/1762758236509_IG_Influencers_100.xlsx"
        ]
        target_sheet = "category ig"
    else:  # TikTok
        possible_names = [
            "Ven_TT_CVI.xlsx",
            "/mnt/user-data/uploads/1762758161982_Ven_TT_CVI.xlsx"
        ]
        target_sheet = "category tt"
    
    # Check each possible location
    for path in possible_names:
        if os.path.exists(path):
            return path, target_sheet
    
    return None, target_sheet

# Title and description
st.title("🔗 Link Batch Processor")
st.markdown("""
Split your link files into batches of 100 and generate multiple template files ready for processing.

**How it works:**
1. Select your platform (Instagram or TikTok)
2. Upload your CSV/Excel file with links
3. Download all batch files as a ZIP
""")

st.divider()

# Platform selection
col1, col2 = st.columns(2)

with col1:
    platform = st.selectbox(
        "Select Platform",
        options=["Instagram", "TikTok"],
        help="Choose the platform for your links"
    )

with col2:
    batch_size = st.number_input(
        "Batch Size",
        min_value=1,
        max_value=500,
        value=100,
        help="Number of links per file (default: 100)"
    )

st.divider()

# Optional: Custom template upload
with st.expander("📤 Advanced: Upload Custom Templates (Optional)"):
    st.markdown("If you want to use different template files, upload them here:")
    
    col1, col2 = st.columns(2)
    with col1:
        ig_template = st.file_uploader("Instagram Template", type=["xlsx"], key="ig_template")
    with col2:
        tt_template = st.file_uploader("TikTok Template", type=["xlsx"], key="tt_template")
    
    # Save uploaded templates temporarily
    if ig_template:
        with open("IG_Influencers_100.xlsx", "wb") as f:
            f.write(ig_template.getbuffer())
        st.success("✅ Instagram template uploaded")
    
    if tt_template:
        with open("Ven_TT_CVI.xlsx", "wb") as f:
            f.write(tt_template.getbuffer())
        st.success("✅ TikTok template uploaded")

st.divider()

# File upload
uploaded_file = st.file_uploader(
    "Upload your link file (CSV or Excel)",
    type=["csv", "xlsx", "xls"],
    help="Upload a CSV or Excel file containing your links"
)

if uploaded_file:
    try:
        # Load the uploaded file
        if uploaded_file.name.endswith('.csv'):
            df = pd.read_csv(uploaded_file)
        else:
            df = pd.read_excel(uploaded_file)
        
        st.success(f"✅ File loaded: {len(df)} rows found")
        
        # Auto-detect link column
        possible_columns = ['Post link', 'post link', 'Post Link', 'URL', 'url', 'Link', 'link']
        link_column = None
        
        for col in possible_columns:
            if col in df.columns:
                link_column = col
                break
        
        if not link_column:
            # Let user select column
            link_column = st.selectbox(
                "Select the column containing links",
                options=df.columns.tolist()
            )
        else:
            st.info(f"📍 Detected link column: **{link_column}**")
        
        # Extract links and remove nulls
        links = df[link_column].dropna().tolist()
        total_links = len(links)
        
        # Calculate batches
        num_files = (total_links + batch_size - 1) // batch_size
        
        # Display stats
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Total Links", f"{total_links:,}")
        with col2:
            st.metric("Batch Size", batch_size)
        with col3:
            st.metric("Output Files", num_files)
        
        st.divider()
        
        # Preview
        with st.expander("📋 Preview First 10 Links"):
            preview_df = pd.DataFrame({link_column: links[:10]})
            st.dataframe(preview_df, use_container_width=True)
        
        st.divider()
        
        # Process button
        if st.button("🚀 Generate Batch Files", type="primary", use_container_width=True):
            
            # Find template file
            template_path, target_sheet = find_template_file(platform)
            
            if not template_path:
                st.error(f"❌ {platform} template file not found. Please ensure template files are in the app directory.")
                st.stop()
            
            with st.spinner(f"Creating {num_files} batch files..."):
                
                # Create ZIP file in memory
                zip_buffer = io.BytesIO()
                
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                    
                    progress_bar = st.progress(0)
                    status_text = st.empty()
                    
                    for i in range(num_files):
                        # Calculate batch
                        start_idx = i * batch_size
                        end_idx = min((i + 1) * batch_size, total_links)
                        batch_links = links[start_idx:end_idx]
                        
                        # Update progress
                        progress = (i + 1) / num_files
                        progress_bar.progress(progress)
                        status_text.text(f"Processing batch {i + 1}/{num_files}...")
                        
                        # Load template
                        wb = load_workbook(template_path)
                        ws = wb[target_sheet]
                        
                        # Clear existing data (keep header)
                        max_row = ws.max_row
                        if max_row > 1:
                            ws.delete_rows(2, max_row - 1)
                        
                        # Insert new links
                        for row_idx, link in enumerate(batch_links, start=2):
                            ws.cell(row=row_idx, column=1, value=link)
                        
                        # Save to buffer
                        excel_buffer = io.BytesIO()
                        wb.save(excel_buffer)
                        excel_buffer.seek(0)
                        
                        # Add to ZIP
                        filename = f"{platform}_Batch_{i+1:02d}_Links_{start_idx+1}-{end_idx}.xlsx"
                        zip_file.writestr(filename, excel_buffer.getvalue())
                    
                    progress_bar.progress(1.0)
                    status_text.text("✅ All files generated!")
                
                # Prepare download
                zip_buffer.seek(0)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                zip_filename = f"{platform}_Batches_{timestamp}.zip"
                
                st.success(f"🎉 Successfully created {num_files} batch files!")
                
                # Download button
                st.download_button(
                    label="📥 Download All Batch Files (ZIP)",
                    data=zip_buffer,
                    file_name=zip_filename,
                    mime="application/zip",
                    use_container_width=True
                )
    
    except Exception as e:
        st.error(f"❌ Error processing file: {str(e)}")
        st.exception(e)

else:
    st.info("👆 Upload a file to get started")

# Footer
st.divider()
st.caption("Built for efficient link batch processing | ConvoSight Analytics")
